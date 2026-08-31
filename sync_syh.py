"""Sincroniza el cronograma de Seguridad e Higiene desde el Google Sheet que
administra la consultora externa (una pestaña por sede/planta). El sheet está
compartido como 'Cualquiera con el enlace, Lector', así que se descarga como
.xlsx con una URL de exportación normal — sin API de Google, sin credenciales.

Cada pestaña tiene un bloque de 3 filas por estudio: la fila del ítem (nombre +
frecuencia), una fila 'Realizado' (con la fecha de la última vez que se hizo,
puesta en algún lugar de la grilla mensual) y una fila 'Informe'. Se recorre
así porque es una planilla armada a mano, no una base de datos: no hay una
columna de 'vencimiento' explícita.

Reimportable: hace upsert por (sede, nombre) — no duplica estudios, actualiza
la última fecha realizada y recalcula el próximo vencimiento cada vez que corre."""
import calendar
import sys
import urllib.request
from datetime import date, datetime

import openpyxl

from config import load_cfg
from db import get_db, init_db

FRECUENCIAS_MESES = {
    "MENSUAL": 1, "BIMESTRAL": 2, "TRIMESTRAL": 3, "CUATRIMESTRAL": 4,
    "SEMESTRAL": 6, "ANUAL": 12, "BIANUAL": 24, "BIENAL": 24,
}

NOMBRES_ESPECIALES = {"REALIZADO", "INFORME"}


def _url_export_xlsx(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def _descargar_xlsx(sheet_id: str, destino):
    url = _url_export_xlsx(sheet_id)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        destino.write_bytes(resp.read())


def _meses_desde_frecuencia(frecuencia):
    texto = str(frecuencia).strip().upper()
    # Orden importante: BIANUAL/BIENAL antes que ANUAL, porque "ANUAL" es substring de "BIANUAL".
    for clave in ("BIANUAL", "BIENAL", "CUATRIMESTRAL", "TRIMESTRAL", "BIMESTRAL", "SEMESTRAL", "MENSUAL", "ANUAL"):
        if clave in texto:
            return FRECUENCIAS_MESES[clave]
    return None


def _proximo_vencimiento(ultima_fecha_str, frecuencia):
    if not ultima_fecha_str or not frecuencia:
        return None
    meses = _meses_desde_frecuencia(frecuencia)
    if not meses:
        return None
    fecha = datetime.strptime(ultima_fecha_str, "%Y-%m-%d").date()
    total = fecha.month - 1 + meses
    anio = fecha.year + total // 12
    mes = total % 12 + 1
    dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
    return date(anio, mes, dia).strftime("%Y-%m-%d")


def _fechas_en_fila(fila):
    return [v for v in fila if isinstance(v, datetime)]


def _procesar_hoja(ws):
    """Devuelve una lista de dicts: sede se agrega afuera. Cada dict tiene
    categoria, nombre, frecuencia, comentario, ultima_fecha_realizado."""
    filas = list(ws.iter_rows(values_only=True))

    fila_encabezado = None
    for i, fila in enumerate(filas):
        textos = [str(v).strip().upper() if v else "" for v in fila]
        if any("MEDICI" in t for t in textos) and any("FRECUENCIA" in t for t in textos):
            fila_encabezado = i
            break
    if fila_encabezado is None:
        return []

    estudios = []
    categoria_actual = None
    item_actual = None

    for fila in filas[fila_encabezado + 1:]:
        col_nombre = fila[1] if len(fila) > 1 else None
        col_frecuencia = fila[2] if len(fila) > 2 else None
        if col_nombre is None or not str(col_nombre).strip():
            continue
        nombre_txt = str(col_nombre).strip()
        nombre_upper = nombre_txt.upper()

        if nombre_upper == "REALIZADO":
            if item_actual is not None:
                fechas = _fechas_en_fila(fila)
                if fechas:
                    item_actual["ultima_fecha_realizado"] = max(fechas).strftime("%Y-%m-%d")
            continue

        if nombre_upper == "INFORME":
            continue

        if isinstance(col_frecuencia, (int, float)):
            # Filas de resumen al pie de la planilla (Estudios planificados, % Cumplimiento,
            # etc.) tienen un número en la columna de frecuencia, no una frecuencia real.
            continue

        if col_frecuencia and str(col_frecuencia).strip():
            item_actual = {
                "categoria": categoria_actual,
                "nombre": nombre_txt,
                "frecuencia": str(col_frecuencia).strip(),
                "comentario": None,
                "ultima_fecha_realizado": None,
            }
            textos_resto = [str(v).strip() for v in fila[3:] if v and isinstance(v, str) and str(v).strip()]
            if textos_resto:
                item_actual["comentario"] = " / ".join(textos_resto)
            estudios.append(item_actual)
        else:
            categoria_actual = nombre_txt
            item_actual = None

    return estudios


def sincronizar(sheet_id: str = None):
    init_db()
    cfg = load_cfg()
    sheet_id = sheet_id or cfg.get("syh_sheet_id")
    if not sheet_id:
        print("Falta 'syh_sheet_id' en rrhh.cfg.")
        return

    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as tmp:
        destino = Path(tmp) / "syh.xlsx"
        _descargar_xlsx(sheet_id, destino)
        wb = openpyxl.load_workbook(destino, data_only=True)

        conn = get_db()
        nuevos = 0
        actualizados = 0

        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            estudios = _procesar_hoja(ws)
            sede = nombre_hoja.strip()

            for est in estudios:
                proximo = _proximo_vencimiento(est["ultima_fecha_realizado"], est["frecuencia"])
                existente = conn.execute(
                    "SELECT id, proximo_vencimiento FROM syh_estudios WHERE sede = ? AND nombre = ?",
                    (sede, est["nombre"]),
                ).fetchone()
                if existente:
                    conn.execute(
                        """UPDATE syh_estudios SET categoria=?, frecuencia=?, comentario=?,
                           ultima_fecha_realizado=?, proximo_vencimiento=?, fecha_sync=datetime('now','localtime')
                           WHERE id=?""",
                        (est["categoria"], est["frecuencia"], est["comentario"],
                         est["ultima_fecha_realizado"], proximo, existente["id"]),
                    )
                    if existente["proximo_vencimiento"] != proximo:
                        actualizados += 1
                else:
                    conn.execute(
                        """INSERT INTO syh_estudios (sede, categoria, nombre, frecuencia, comentario,
                           ultima_fecha_realizado, proximo_vencimiento)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (sede, est["categoria"], est["nombre"], est["frecuencia"], est["comentario"],
                         est["ultima_fecha_realizado"], proximo),
                    )
                    nuevos += 1

        conn.commit()
        conn.close()
        print(f"Seguridad e Higiene: {nuevos} estudios nuevos, {actualizados} con cambios.")


if __name__ == "__main__":
    sincronizar(sys.argv[1] if len(sys.argv) > 1 else None)
