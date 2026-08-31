"""Importa el historial de capacitaciones desde 'Capacitaciones 2026.xlsx':
cada hoja del archivo es una capacitación distinta (nombre de la hoja = tema),
con la fecha de dictado en algún lugar de la fila de encabezado y, debajo,
una fila por persona que la tomó (DNI, nombre, puntaje opcional). Se cruza
por DNI contra empleados; si no matchea, se guarda igual con el nombre suelto.
Reimportable: se dedupe por (tema, nombre, fecha)."""
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from db import get_db, init_db

EXCEL_DIR = Path(__file__).resolve().parent / "excel"

COLUMNAS_PUNTUACION = {"puntuacion", "puntuación"}
COLUMNAS_DNI = {"dni", "dni (sin punto)"}
COLUMNAS_NOMBRE = {"nombre y apellido", "nombre", "apellido y nombre"}

# El "Tema madre" de cada capacitación está codificado como color de la pestaña
# de la hoja en el Excel (no hay una columna de texto para esto). Mapeo confirmado
# con Oscar: verde = Seguridad e Higiene, naranja = Medicina Laboral, violeta = Medio Ambiente.
TEMA_MADRE_POR_COLOR = {
    "93C47D": "Seguridad e Higiene",
    "FF9900": "Medicina Laboral",
    "5B3F86": "Medio Ambiente",
}


def _tema_madre_por_color_pestana(ws):
    color = ws.sheet_properties.tabColor
    if not color or not color.rgb:
        return None
    rgb = str(color.rgb)[-6:].upper()
    return TEMA_MADRE_POR_COLOR.get(rgb)


def _normalizar_encabezado(valor):
    return str(valor).strip().lower() if valor is not None else ""


def _normalizar_dni(valor):
    if valor is None:
        return None
    if isinstance(valor, float):
        valor = int(valor)
    digitos = re.sub(r"\D", "", str(valor))
    return digitos or None


def _mapear_columnas(fila_encabezado):
    col_puntuacion = col_dni = col_nombre = None
    fecha_sesion = None
    for idx, valor in enumerate(fila_encabezado):
        encabezado = _normalizar_encabezado(valor)
        if encabezado in COLUMNAS_PUNTUACION:
            col_puntuacion = idx
        elif encabezado in COLUMNAS_DNI:
            col_dni = idx
        elif encabezado in COLUMNAS_NOMBRE:
            col_nombre = idx
        if isinstance(valor, datetime):
            fecha_sesion = valor
    return col_puntuacion, col_dni, col_nombre, fecha_sesion


def _obtener_o_crear_tema(conn, nombre, tema_madre=None):
    fila = conn.execute("SELECT id, tema_madre FROM capacitaciones_temas WHERE nombre = ?", (nombre,)).fetchone()
    if fila:
        if tema_madre and not fila["tema_madre"]:
            conn.execute("UPDATE capacitaciones_temas SET tema_madre = ? WHERE id = ?", (tema_madre, fila["id"]))
        return fila["id"]
    return conn.execute(
        "INSERT INTO capacitaciones_temas (nombre, tipo, tema_madre) VALUES (?, 'Capacitación', ?)",
        (nombre, tema_madre),
    ).lastrowid


def importar(ruta_excel: Path):
    init_db()
    conn = get_db()

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    nuevos = 0
    duplicados = 0
    sin_fecha_hojas = []
    sin_empleado = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue
        col_puntuacion, col_dni, col_nombre, fecha_sesion = _mapear_columnas(filas[0])
        if col_nombre is None:
            print(f"[{nombre_hoja}] no encontré columna de nombre, la salteo.")
            continue
        if fecha_sesion is None:
            fecha_sesion_str = "2026-01-01"
            sin_fecha_hojas.append(nombre_hoja)
        else:
            fecha_sesion_str = fecha_sesion.strftime("%Y-%m-%d")

        tema_madre = _tema_madre_por_color_pestana(ws)
        tema_id = _obtener_o_crear_tema(conn, nombre_hoja.strip(), tema_madre)

        for fila in filas[1:]:
            if col_nombre >= len(fila):
                continue
            nombre = fila[col_nombre]
            if not nombre or not str(nombre).strip():
                continue
            nombre = str(nombre).strip()
            if nombre.upper() in {"TOTAL", "TOTALES"}:
                continue

            dni = _normalizar_dni(fila[col_dni]) if col_dni is not None and col_dni < len(fila) else None
            puntuacion = None
            if col_puntuacion is not None and col_puntuacion < len(fila):
                val = fila[col_puntuacion]
                if val is not None and str(val).strip() != "":
                    try:
                        puntuacion = float(val)
                    except (TypeError, ValueError):
                        puntuacion = None

            empleado = None
            if dni:
                empleado = conn.execute(
                    "SELECT id FROM empleados WHERE REPLACE(REPLACE(dni,'.',''),'-','') = ?", (dni,)
                ).fetchone()
            if not empleado:
                empleado = conn.execute(
                    "SELECT id FROM empleados WHERE UPPER(TRIM(nombre)) = ?", (nombre.upper(),)
                ).fetchone()
            empleado_id = empleado["id"] if empleado else None
            if not empleado_id:
                sin_empleado.append(nombre)

            cur = conn.execute(
                """INSERT OR IGNORE INTO capacitaciones_registros
                   (tema_id, empleado_id, nombre_original, fecha, puntuacion)
                   VALUES (?, ?, ?, ?, ?)""",
                (tema_id, empleado_id, nombre, fecha_sesion_str, puntuacion),
            )
            if cur.rowcount:
                nuevos += 1
            else:
                duplicados += 1

    conn.commit()
    conn.close()
    print(f"Capacitaciones: {nuevos} registros nuevos, {duplicados} ya existentes (sin duplicar).")
    if sin_fecha_hojas:
        print(f"Hojas sin fecha en el encabezado (se les puso 2026-01-01, revisar a mano): {sin_fecha_hojas}")
    if sin_empleado:
        print(f"Sin match de empleado ({len(set(sin_empleado))} nombres distintos): {sorted(set(sin_empleado))}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        ruta = Path(sys.argv[1])
    else:
        candidatos = sorted(EXCEL_DIR.glob("Capacitaciones*.xls*"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidatos:
            print(f"No encontré ningún archivo 'Capacitaciones*.xls*' en {EXCEL_DIR}")
            sys.exit(1)
        ruta = candidatos[0]
    print(f"Importando desde: {ruta}")
    importar(ruta)
